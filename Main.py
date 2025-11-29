import logging
import sqlite3
import os
from datetime import datetime
from dataclasses import dataclass
from typing import List, Dict, Optional
import openpyxl
from openpyxl import Workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler, 
    filters, ContextTypes, ConversationHandler
)

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

GET_NAME, MAIN_MENU = range(2)

@dataclass
class Product:
    id: int
    name: str
    price: float
    description: str

class ProductRepository:
    def __init__(self, filename: str = "products.txt"):
        self.filename = filename
        self._ensure_products_file()
    
    def _ensure_products_file(self):
        if not os.path.exists(self.filename):
            default_products = [
                "Пластик PLA|150.00|Качественный PLA пластик для 3D печати",
                "Пластик ABS|180.00|Прочный ABS пластик",
                "Пластик PETG|200.00|Гибкий PETG пластик",
                "Подставка для телефона|300.00|Стильная подставка для смартфона",
                "Чехол для наушников|250.00|Защитный чехол для беспроводных наушников",
                "Статуэтка персонажа|500.00|Кастомная фигурка по вашему дизайну"
            ]
            with open(self.filename, 'w', encoding='utf-8') as f:
                f.write('\n'.join(default_products))
            logger.info("Создан файл с товарами по умолчанию")
    
    def get_all_products(self) -> List[Product]:
        products = []
        try:
            with open(self.filename, 'r', encoding='utf-8') as f:
                for idx, line in enumerate(f, 1):
                    line = line.strip()
                    if line and '|' in line:
                        name, price, description = line.split('|', 2)
                        products.append(Product(
                            id=idx,
                            name=name.strip(),
                            price=float(price.strip()),
                            description=description.strip()
                        ))
        except Exception as e:
            logger.error(f"Ошибка чтения товаров: {e}")
        return products
    
    def get_product_by_id(self, product_id: int) -> Optional[Product]:
        products = self.get_all_products()
        return next((p for p in products if p.id == product_id), None)

class DatabaseManager:
    def __init__(self, db_name: str = "print_shop.db"):
        self.db_name = db_name
        self.product_repo = ProductRepository()
        self._init_db()
    
    def _init_db(self):
        with sqlite3.connect(self.db_name) as conn:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    user_id INTEGER PRIMARY KEY,
                    first_name TEXT NOT NULL,
                    last_name TEXT NOT NULL,
                    registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            conn.execute('''
                CREATE TABLE IF NOT EXISTS cart_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    quantity INTEGER NOT NULL,
                    added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users (user_id)
                )
            ''')
    
    def add_user(self, user_id: int, first_name: str, last_name: str):
        with sqlite3.connect(self.db_name) as conn:
            conn.execute(
                'INSERT OR REPLACE INTO users (user_id, first_name, last_name) VALUES (?, ?, ?)',
                (user_id, first_name, last_name)
            )
    
    def get_user(self, user_id: int) -> Optional[tuple]:
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.execute(
                'SELECT first_name, last_name FROM users WHERE user_id = ?',
                (user_id,)
            )
            return cursor.fetchone()
    
    def add_to_cart(self, user_id: int, product_id: int, quantity: int):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.execute(
                'SELECT quantity FROM cart_items WHERE user_id = ? AND product_id = ?',
                (user_id, product_id)
            )
            existing = cursor.fetchone()
            
            if existing:
                new_quantity = existing[0] + quantity
                conn.execute(
                    'UPDATE cart_items SET quantity = ? WHERE user_id = ? AND product_id = ?',
                    (new_quantity, user_id, product_id)
                )
            else:
                conn.execute(
                    'INSERT INTO cart_items (user_id, product_id, quantity) VALUES (?, ?, ?)',
                    (user_id, product_id, quantity)
                )
    
    def get_cart_items(self, user_id: int) -> List[tuple]:
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.execute(
                'SELECT product_id, quantity FROM cart_items WHERE user_id = ?',
                (user_id,)
            )
            cart_data = cursor.fetchall()
            
        products = self.product_repo.get_all_products()
        result = []
        
        for product_id, quantity in cart_data:
            product = next((p for p in products if p.id == product_id), None)
            if product:
                result.append((product.id, product.name, product.price, quantity))
        
        return result
    
    def clear_cart(self, user_id: int):
        with sqlite3.connect(self.db_name) as conn:
            conn.execute('DELETE FROM cart_items WHERE user_id = ?', (user_id,))

class ExcelExporter:
    def __init__(self, filename: str = "orders.xlsx"):
        self.filename = filename
        self._ensure_excel_file()
    
    def _ensure_excel_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Заказы"
            headers = ["Дата", "Имя Фамилия", "Товар 1", "Кол-во 1", "Товар 2", "Кол-во 2", 
                      "Товар 3", "Кол-во 3", "Товар 4", "Кол-во 4"]
            ws.append(headers)
            wb.save(self.filename)
            logger.info("Создан Excel файл для заказов")
    
    def add_order(self, full_name: str, cart_items: List[tuple]):
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            row_data = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                full_name
            ]
            
            for item in cart_items:
                product_id, name, price, quantity = item
                row_data.extend([name, quantity])
            
            ws.append(row_data)
            wb.save(self.filename)
            logger.info(f"Добавлен заказ для {full_name}")
            return True
        except Exception as e:
            logger.error(f"Ошибка записи в Excel: {e}")
            return False
    
    def get_user_orders(self, full_name: str) -> List[dict]:
        try:
            if not os.path.exists(self.filename):
                return []
            
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            orders = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue
                    
                if row[1] == full_name: 
                    order_data = {
                        'date': row[0],
                        'items': []
                    }
                    
                    for i in range(2, len(row), 2):
                        if i < len(row) and row[i] and row[i+1]:
                            order_data['items'].append({
                                'product': row[i],
                                'quantity': row[i+1]
                            })
                    
                    orders.append(order_data)
            
            return orders
        except Exception as e:
            logger.error(f"Ошибка чтения заказов из Excel: {e}")
            return []

db = DatabaseManager()
product_repo = ProductRepository()
exporter = ExcelExporter()

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    logger.info(f"Получена команда /start от пользователя {user_id}")
    
    user_data = db.get_user(user_id)
    
    if user_data:
        first_name, last_name = user_data
        await show_main_menu(update, context, first_name, last_name)
        return MAIN_MENU
    else:
        await update.message.reply_text(
            "👋 Привет! Я бот для заказа 3D печати!\n"
            "Для начала расскажи немного о себе:\n"
            "Введи своё Имя и Фамилию (например: Иван Иванов)"
        )
        return GET_NAME

async def get_user_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    full_name = update.message.text.strip()
    logger.info(f"Пользователь {user_id} ввел имя: {full_name}")
    
    name_parts = full_name.split()
    if len(name_parts) < 2:
        await update.message.reply_text(
            "❌ Пожалуйста, введите Имя и Фамилию через пробел (например: Иван Иванов)"
        )
        return GET_NAME
    
    first_name = name_parts[0]
    last_name = ' '.join(name_parts[1:])
    
    db.add_user(user_id, first_name, last_name)
    
    await update.message.reply_text(
        f"✅ Отлично, {first_name}! Регистрация завершена!\n"
        f"Теперь ты можешь заказывать товары для 3D печати 🎨"
    )
    
    await show_main_menu(update, context, first_name, last_name)
    return MAIN_MENU

async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, first_name: str, last_name: str):
    keyboard = [
        [InlineKeyboardButton("📦 Каталог товаров", callback_data="catalog")],
        [InlineKeyboardButton("🛒 Корзина", callback_data="cart")],
        [InlineKeyboardButton("📋 Мои заказы", callback_data="orders")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = f"👋 {first_name}, добро пожаловать в магазин 3D печати!\nВыберите действие:"
    
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text, reply_markup=reply_markup)

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_data = db.get_user(query.from_user.id)
    if not user_data:
        await query.edit_message_text("❌ Ошибка: пользователь не найден")
        return
    
    first_name, last_name = user_data
    
    if query.data == "catalog":
        await show_catalog(update, context)
    elif query.data == "cart":
        await show_cart(update, context)
    elif query.data == "orders":
        await show_user_orders(update, context, first_name, last_name)
    elif query.data == "back_to_menu":
        await show_main_menu(update, context, first_name, last_name)
    elif query.data.startswith("product_"):
        await show_product_details(update, context)
    elif query.data.startswith("qty_"):
        await add_to_cart_handler(update, context)
    elif query.data == "checkout":
        await checkout(update, context)
    elif query.data == "clear_cart":
        await clear_cart(update, context)

async def show_catalog(update: Update, context: ContextTypes.DEFAULT_TYPE):
    products = product_repo.get_all_products()
    
    if not products:
        await update.callback_query.edit_message_text("😔 Каталог товаров пуст")
        return
    
    keyboard = []
    for product in products:
        keyboard.append([
            InlineKeyboardButton(
                f"{product.name} - {product.price}₽", 
                callback_data=f"product_{product.id}"
            )
        ])
    
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="back_to_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = "🛍️ **Каталог товаров:**\n\n"
    for product in products:
        text += f"• {product.name} - {product.price}₽\n  {product.description}\n\n"
    
    await update.callback_query.edit_message_text(
        text, 
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def show_product_details(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    product_id = int(query.data.split('_')[1])
    
    product = product_repo.get_product_by_id(product_id)
    if not product:
        await query.edit_message_text("❌ Товар не найден")
        return
    
    context.user_data['selected_product'] = product_id
    
    keyboard = [
        [InlineKeyboardButton("1", callback_data="qty_1"),
         InlineKeyboardButton("2", callback_data="qty_2"),
         InlineKeyboardButton("3", callback_data="qty_3")],
        [InlineKeyboardButton("5", callback_data="qty_5"),
         InlineKeyboardButton("10", callback_data="qty_10")],
        [InlineKeyboardButton("🔙 Назад", callback_data="catalog")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = (
        f"🎯 **{product.name}**\n\n"
        f"📝 {product.description}\n"
        f"💰 Цена: {product.price}₽ за шт.\n\n"
        f"Выберите количество:"
    )
    
    await query.edit_message_text(
        text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def add_to_cart_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    quantity = int(query.data.split('_')[1])
    product_id = context.user_data.get('selected_product')
    user_id = query.from_user.id
    
    if not product_id:
        await query.edit_message_text("❌ Ошибка: товар не выбран")
        return
    
    product = product_repo.get_product_by_id(product_id)
    if not product:
        await query.edit_message_text("❌ Товар не найден")
        return

    db.add_to_cart(user_id, product_id, quantity)
    
    await query.edit_message_text(
        f"✅ {product.name} x{quantity} добавлен в корзину!\n\n"
        f"Что хотите сделать дальше?",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📦 Продолжить покупки", callback_data="catalog")],
            [InlineKeyboardButton("🛒 Перейти в корзину", callback_data="cart")],
            [InlineKeyboardButton("🔙 В меню", callback_data="back_to_menu")]
        ])
    )

async def show_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.callback_query.from_user.id
    cart_items = db.get_cart_items(user_id)
    
    if not cart_items:
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="back_to_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.callback_query.edit_message_text(
            "🛒 Ваша корзина пуста",
            reply_markup=reply_markup
        )
        return
    
    total = 0
    text = "🛒 **Ваша корзина:**\n\n"
    for item in cart_items:
        product_id, name, price, quantity = item
        item_total = price * quantity
        total += item_total
        text += f"• {name} x{quantity} = {item_total}₽\n"
    
    text += f"\n💵 **Итого: {total}₽**"
    
    keyboard = [
        [InlineKeyboardButton("✅ Оформить заказ", callback_data="checkout")],
        [InlineKeyboardButton("🗑️ Очистить корзину", callback_data="clear_cart")],
        [InlineKeyboardButton("🔙 Назад", callback_data="back_to_menu")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.callback_query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')

async def show_user_orders(update: Update, context: ContextTypes.DEFAULT_TYPE, first_name: str, last_name: str):
    full_name = f"{first_name} {last_name}"
    orders = exporter.get_user_orders(full_name)
    
    if not orders:
        await update.callback_query.edit_message_text(
            "📋 **История заказов**\n\n"
            "У вас еще нет завершенных заказов.\n"
            "Сделайте свой первый заказ в разделе 📦 Каталог товаров!",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📦 Перейти в каталог", callback_data="catalog")],
                [InlineKeyboardButton("🔙 В меню", callback_data="back_to_menu")]
            ]),
            parse_mode='Markdown'
        )
        return
    
    page = context.user_data.get('orders_page', 0)
    orders_per_page = 3
    total_pages = (len(orders) + orders_per_page - 1)
    orders_per_page
    start_idx = page * orders_per_page
    end_idx = start_idx + orders_per_page
    current_orders = orders[start_idx:end_idx]
    
    text = f"📋 **История заказов**\n\n"
    text += f"Всего заказов: {len(orders)}\n\n"
    
    for i, order in enumerate(current_orders, start_idx + 1):
        text += f"**Заказ #{i}** - {order['date']}\n"
        
        total_amount = 0
        for item in order['items']:
            product = next((p for p in product_repo.get_all_products() if p.name == item['product']), None)
            if product:
                item_total = product.price * item['quantity']
                total_amount += item_total
                text += f"  • {item['product']} x{item['quantity']} = {item_total}₽\n"
            else:
                text += f"  • {item['product']} x{item['quantity']}\n"
        
        text += f"  **Итого: {total_amount}₽**\n\n"
    
    keyboard = []

    pagination_buttons = []
    if page > 0:
        pagination_buttons.append(InlineKeyboardButton("⬅️ Назад", callback_data=f"orders_page_{page-1}"))
    
    pagination_buttons.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="current_page"))
    
    if page < total_pages - 1:
        pagination_buttons.append(InlineKeyboardButton("Вперед ➡️", callback_data=f"orders_page_{page+1}"))
    
    if pagination_buttons:
        keyboard.append(pagination_buttons)
    
    keyboard.extend([
        [InlineKeyboardButton("📦 Новый заказ", callback_data="catalog")],
        [InlineKeyboardButton("🔙 В меню", callback_data="back_to_menu")]
    ])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.callback_query.edit_message_text(
        text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def checkout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.callback_query.from_user.id
    user_data = db.get_user(user_id)
    
    if not user_data:
        await update.callback_query.edit_message_text("❌ Ошибка: пользователь не найден")
        return
    
    first_name, last_name = user_data
    full_name = f"{first_name} {last_name}"
    cart_items = db.get_cart_items(user_id)
    
    if not cart_items:
        await update.callback_query.edit_message_text("❌ Корзина пуста")
        return
    
    success = exporter.add_order(full_name, cart_items)
    
    if success:
        db.clear_cart(user_id)

        order_text = "✅ **Заказ оформлен!**\n\nВаш заказ:\n"
        total = 0
        for item in cart_items:
            product_id, name, price, quantity = item
            item_total = price * quantity
            total += item_total
            order_text += f"• {name} x{quantity} = {item_total}₽\n"
        
        order_text += f"\n💵 **Общая сумма: {total}₽**\n\n"
        order_text += "📋 Заказ записан в таблицу Excel. Спасибо!"
        
        keyboard = [
            [InlineKeyboardButton("📋 Посмотреть заказы", callback_data="orders")],
            [InlineKeyboardButton("🔙 В меню", callback_data="back_to_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.callback_query.edit_message_text(
            order_text,
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
    else:
        await update.callback_query.edit_message_text(
            "❌ Ошибка при оформлении заказа\n\nПопробуйте позже или обратитесь к администратору.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Назад", callback_data="back_to_menu")]
            ])
        )

async def clear_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.callback_query.from_user.id
    db.clear_cart(user_id)
    
    await update.callback_query.edit_message_text(
        "🗑️ Корзина очищена!",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔙 В меню", callback_data="back_to_menu")]
        ])
    )

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Операция отменена")
    return ConversationHandler.END

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.error(f"Ошибка при обработке update {update}: {context.error}")
    
    if update and update.effective_message:
        await update.effective_message.reply_text(
            "❌ Произошла ошибка при обработке запроса.\n\nПопробуйте еще раз или обратитесь к администратору."
        )

def main():
    token = "8407196376:AAH2rNvjqkbBb-ndbo8BmUlHlx6nLdseL48"
    
    if not token:
        logger.error("❌ ТОКЕН БОТА НЕ НАСТРОЕН!")
        return
    
    application = Application.builder().token(token).build()
    
    application.add_error_handler(error_handler)
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            GET_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, get_user_name)
            ],
            MAIN_MENU: [
                CallbackQueryHandler(handle_callback)
            ],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    application.add_handler(conv_handler)
    
    application.add_handler(CallbackQueryHandler(handle_callback, pattern="^(catalog|cart|orders|back_to_menu|product_|qty_|checkout|clear_cart)$"))
    
    application.add_handler(CallbackQueryHandler(handle_callback, pattern="^orders_page_"))
    
    logger.info("🤖 Бот запускается...")
    application.run_polling()

if __name__ == '__main__':
    main()